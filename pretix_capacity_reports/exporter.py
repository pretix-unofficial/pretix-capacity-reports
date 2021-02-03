import json
import tempfile
from collections import OrderedDict
from datetime import timedelta, time, datetime

import pytz
from dateutil.parser import parse
from django import forms
from django.db.models import OuterRef, Exists, Prefetch, Count, Q, Sum, Min, Subquery, Max, F, CharField, Value
from django.db.models.functions import TruncDay, Coalesce, Cast, Concat
from django.utils.functional import cached_property
from django.utils.timezone import now, get_current_timezone, make_aware
from django.utils.translation import gettext_lazy as _
from i18nfield.strings import LazyI18nString
from openpyxl import Workbook
from openpyxl.cell.cell import KNOWN_TYPES
from openpyxl.utils import get_column_letter

from pretix.base.exporter import MultiSheetListExporter
from pretix.base.models import Quota, EventMetaValue, Order, OrderPosition, SubEvent, Checkin, LogEntry, Item, ItemVariation


class BaseMSLE(MultiSheetListExporter):
    def _render_xlsx(self, form_data, output_file=None):  # vendored pretix 3.16 version
        wb = Workbook(write_only=True)
        n_sheets = len(self.sheets)
        for i_sheet, (s, l) in enumerate(self.sheets):
            ws = wb.create_sheet(str(l))
            if hasattr(self, 'prepare_xlsx_sheet_' + s):
                getattr(self, 'prepare_xlsx_sheet_' + s)(ws)

            total = 0
            counter = 0
            for i, line in enumerate(self.iterate_sheet(form_data, sheet=s)):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                ws.append([
                    str(val) if not isinstance(val, KNOWN_TYPES) else val
                    for val in line
                ])
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100 / n_sheets + 100 / n_sheets * i_sheet)

        if output_file:
            wb.save(output_file)
            return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', None
        else:
            with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
                wb.save(f.name)
                f.seek(0)
                return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f.read()

    @cached_property
    def product_choices(self):
        product_choices = [('', 'All')]
        for r in Item.objects.filter(
            event__in=self.events, variations__isnull=True,
        ).annotate(
            n=Cast('name', output_field=CharField())
        ).values('n').distinct():
            product_choices.append((str(r['n']) + '#!#-', i18ntostring(r['n'])))
        for r in ItemVariation.objects.filter(
                item__event__in=self.events
        ).annotate(
            n=Cast('item__name', output_field=CharField()),
            v=Cast('value', output_field=CharField())
        ).values(
            'n', 'v'
        ).distinct():
            product_choices.append(('{}#!#{}'.format(r['n'], r['v']), '{} – {}'.format(
                i18ntostring(r['n']),
                i18ntostring(r['v']),
            )))

        product_choices.sort(key=lambda i: str(i[1]))
        return product_choices


def i18ntostring(v):
    if v.startswith('{'):
        return LazyI18nString(json.loads(v))
    return v


class CapacityUtilizationReport(BaseMSLE):
    identifier = 'capacity_utilization'
    verbose_name = 'Capacity & Utilization'
    meta_name = 'AgencyNumber'

    sheets = [
        ('date_agency_event', _('By date, agency, and event')),
        ('date_agency', _('By date and agency')),
        ('agency_date_day', _('By agency and day')),
        ('agency_date_week', _('By agency and week')),
    ]

    @property
    def export_form_fields(self):
        defdate_start = now().astimezone(get_current_timezone()).date()
        defdate_end = now().astimezone(get_current_timezone()).date() + timedelta(days=6)

        f = OrderedDict(
            list(super().export_form_fields.items()) + [
                ('date_from',
                 forms.DateField(
                     label=_('Start date'),
                     widget=forms.DateInput(attrs={'class': 'datepickerfield'}),
                     initial=defdate_start,
                 )),
                ('date_to',
                 forms.DateField(
                     label=_('End date'),
                     widget=forms.DateInput(attrs={'class': 'datepickerfield'}),
                     initial=defdate_end,
                 )),
                ('product_name',
                 forms.ChoiceField(
                     label=_('Product and variation'),
                     choices=self.product_choices,
                     required=False
                 )),
            ]
        )
        if self.is_multievent and self.events.first():
            organizer = self.events.first().organizer
            for mp in organizer.meta_properties.prefetch_related('event_values'):
                if mp.name != self.meta_name:
                    continue
                values = sorted(list({v.value for v in mp.event_values.all()}))
                f['meta:{}'.format(mp.name)] = forms.MultipleChoiceField(
                    label=mp.name,
                    choices=[(v, v) for v in values],
                    widget=forms.CheckboxSelectMultiple(
                        attrs={'class': 'scrolling-multiple-choice'}
                    ),
                    initial=values,
                )
        return f

    def iterate_sheet(self, form_data, sheet):
        if self.events.first():
            self.tz = self.events.first().timezone
        else:
            self.tz = pytz.UTC
        self.date_from = form_data['date_from']
        self.date_until = form_data['date_to']
        if isinstance(self.date_from, str):
            self.date_from = parse(self.date_from).date()
        self.datetime_from = make_aware(datetime.combine(
            self.date_from,
            time(hour=0, minute=0, second=0, microsecond=0)
        ), self.tz)

        if isinstance(self.date_until, str):
            self.date_until = parse(self.date_until).date()
        self.datetime_until = make_aware(datetime.combine(
            self.date_until + timedelta(days=1),
            time(hour=0, minute=0, second=0, microsecond=0)
        ), self.tz)

        self.cached_events = list(
            self.events.prefetch_related(
                'organizer', '_settings_objects', 'organizer___settings_objects', 'organizer__meta_properties',
                Prefetch(
                    'meta_values',
                    EventMetaValue.objects.select_related('property'),
                    to_attr='meta_values_cached'
                )
            )
        )

        subevs = {
            (r['day'].astimezone(self.tz).date(), r['event_id']): r['c'] for r in
            self._subevent_qs(form_data).annotate(
                day=TruncDay('date_from', tzinfo=self.tz)
            ).order_by().values('day', 'event_id').annotate(c=Count('*'))
        }
        quotas = {
            (r['day'].astimezone(self.tz).date(), r['event_id']): r['c'] for r in
            self._base_quota_qs(form_data).filter(
                size__isnull=False,
            ).annotate(
                day=TruncDay(Coalesce('subevent__date_from', 'event__date_from'), tzinfo=self.tz)
            ).order_by().values('day', 'event_id').annotate(c=Sum('size'))
        }
        orders = {
            (r['day'].astimezone(self.tz).date(), r['order__event_id']): r['c'] for r in
            self._base_position_qs(form_data).annotate(
                day=TruncDay(Coalesce('subevent__date_from', 'order__event__date_from'), tzinfo=self.tz)
            ).order_by().values('day', 'order__event_id').annotate(c=Count('*'))
        }
        checkins = {
            (r['day'].astimezone(self.tz).date(), r['order__event_id']): r['c'] for r in
            self._base_position_qs(form_data, has_checkin=True).annotate(
                day=TruncDay(Coalesce('subevent__date_from', 'order__event__date_from'), tzinfo=self.tz)
            ).order_by().values('day', 'order__event_id').annotate(c=Count('*'))
        }

        meta_values = form_data['meta:{}'.format(self.meta_name)] if self.is_multievent else [self.event.meta_data[self.meta_name]]
        if hasattr(self, 'iterate_' + sheet):
            yield from getattr(self, 'iterate_' + sheet)(form_data, meta_values, subevs, quotas, orders, checkins)

    def _base_quota_qs(self, form_data):
        qs = Quota.objects.filter(
            Q(subevent__date_from__gte=self.datetime_from, subevent__date_from__lt=self.datetime_until) | Q(subevent__isnull=True, event__date_from__gte=self.datetime_from, event__date_from__lt=self.datetime_until),
            event__in=self.events,
            subevent__date_from__gte=self.datetime_from,
            subevent__date_from__lt=self.datetime_until,
        )
        if form_data['product_name']:
            qs = qs.annotate(
                has_p=Exists(
                    Quota.items.through.objects.annotate(
                        n=Concat(Cast('item__name', output_field=CharField()), Value('#!#-')),
                    ).filter(
                        item__variations__isnull=True,
                        quota=OuterRef('pk'),
                        n=form_data['product_name']
                    )
                ),
                has_v=Exists(
                    Quota.variations.through.objects.annotate(
                        n=Concat(
                            Cast('itemvariation__item__name', output_field=CharField()),
                            Value('#!#'),
                            Cast('itemvariation__value', output_field=CharField()),
                        )
                    ).filter(
                        quota=OuterRef('pk'),
                        n=form_data['product_name']
                    )
                ),
            ).filter(
                Q(has_p=True) | Q(has_v=True)
            )

        for i, n in enumerate([self.meta_name]):
            if 'meta:{}'.format(n) in form_data:
                emv_with_value = EventMetaValue.objects.filter(
                    event=OuterRef('event_id'),
                    property__name=n,
                    value__in=form_data['meta:{}'.format(n)]
                )
                qs = qs.annotate(**{
                    'attr_{}'.format(i): Exists(emv_with_value)
                }).filter(**{
                    'attr_{}'.format(i): True
                })
        # item, variation
        return qs

    def _base_position_qs(self, form_data, has_checkin=False):
        qs = OrderPosition.objects.filter(
            Q(subevent__date_from__gte=self.datetime_from, subevent__date_from__lt=self.datetime_until) | Q(subevent__isnull=True, order__event__date_from__gte=self.datetime_from, order__event__date_from__lt=self.datetime_until),
            order__event__in=self.events,
            order__status__in=(Order.STATUS_PAID, Order.STATUS_PENDING),
        )
        for i, n in enumerate([self.meta_name]):
            if 'meta:{}'.format(n) in form_data:
                emv_with_value = EventMetaValue.objects.filter(
                    event=OuterRef('order__event_id'),
                    property__name=n,
                    value__in=form_data['meta:{}'.format(n)]
                )
                qs = qs.annotate(**{
                    'attr_{}'.format(i): Exists(emv_with_value)
                }).filter(**{
                    'attr_{}'.format(i): True
                })
        if has_checkin:
            qs = qs.annotate(has_checkin=Exists(Checkin.objects.filter(position=OuterRef('pk')))).filter(has_checkin=True)

        if form_data['product_name']:
            qs = qs.annotate(
                n=Concat(
                    Cast('item__name', output_field=CharField()),
                    Value('#!#'),
                    Coalesce(Cast('variation__value', output_field=CharField()), Value('-'))
                )
            ).filter(
                Q(n=form_data['product_name'])
            )

        return qs

    def _subevent_qs(self, form_data):
        qs = SubEvent.objects.filter(
            event__in=self.events,
            date_from__gte=self.datetime_from,
            date_from__lt=self.datetime_until
        )
        for i, n in enumerate([self.meta_name]):
            if 'meta:{}'.format(n) in form_data:
                emv_with_value = EventMetaValue.objects.filter(
                    event=OuterRef('event_id'),
                    property__name=n,
                    value__in=form_data['meta:{}'.format(n)]
                )
                qs = qs.annotate(**{
                    'attr_{}'.format(i): Exists(emv_with_value)
                }).filter(**{
                    'attr_{}'.format(i): True
                })

        if form_data['product_name']:
            qs = qs.annotate(
                has_p=Exists(
                    Quota.items.through.objects.annotate(
                        n=Concat(Cast('item__name', output_field=CharField()), Value('#!#-')),
                    ).filter(
                        item__variations__isnull=True,
                        quota__subevent=OuterRef('pk'),
                        n=form_data['product_name']
                    )
                ),
                has_v=Exists(
                    Quota.variations.through.objects.annotate(
                        n=Concat(
                            Cast('itemvariation__item__name', output_field=CharField()),
                            Value('#!#'),
                            Cast('itemvariation__value', output_field=CharField()),
                        )
                    ).filter(
                        quota__subevent=OuterRef('pk'),
                        n=form_data['product_name']
                    )
                ),
            ).filter(
                Q(has_p=True) | Q(has_v=True)
            )
        return qs

    def _date_iter(self):
        dt = self.date_from
        while dt <= self.date_until:
            yield dt
            dt += timedelta(days=1)

    def _week_iter(self):
        dt = self.date_from
        current_week = []
        while dt <= self.date_until:
            current_week.append(dt)
            if dt.weekday() == 5:  # saturday
                yield current_week
                current_week = []
            dt += timedelta(days=1)
        if current_week:
            yield current_week

    def iterate_date_agency_event(self, form_data, meta_values, subevs, quotas, orders, checkins):
        yield [
            "Date of Event", self.meta_name, "Event ID", "Number of Timeslots", "Sum of Quota", "Sum of Orders", "Sum of Checked in"
        ]

        for mv in meta_values:
            events = sorted([e for e in self.cached_events if e.meta_data[self.meta_name] == mv], key=lambda e: str(e.name))
            for e in events:
                for dt in self._date_iter():
                    subevcnt = subevs.get((dt, e.pk), 0)
                    if e.has_subevents and not subevcnt:
                        continue
                    yield [
                        dt.strftime('%m/%d/%Y'),
                        mv,
                        e.slug,
                        subevcnt,
                        quotas.get((dt, e.pk), 0) or 0,
                        orders.get((dt, e.pk), 0),
                        checkins.get((dt, e.pk), 0),
                    ]

    def prepare_xlsx_sheet_date_agency_event(self, ws):
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

    def iterate_date_agency(self, form_data, meta_values, subevs, quotas, orders, checkins):
        yield [
            "Date of Event", self.meta_name, "Number of Events", "Sum of Quotas", "Sum of Orders", "Sum of Checked in"
        ]

        for mv in meta_values:
            events = sorted([e for e in self.cached_events if e.meta_data[self.meta_name] == mv], key=lambda e: str(e.name))
            for dt in self._date_iter():
                evcnt = sum((1 if not e.has_subevents or subevs.get((dt, e.pk), 0) else 0 for e in events), start=0)
                if not evcnt:
                    continue
                yield [
                    dt.strftime('%m/%d/%Y'),
                    mv,
                    evcnt,
                    sum((quotas.get((dt, e.pk), 0) or 0 for e in events), start=0),
                    sum((orders.get((dt, e.pk), 0) for e in events), start=0),
                    sum((checkins.get((dt, e.pk), 0) for e in events), start=0),
                ]

    def prepare_xlsx_sheet_date_agency(self, ws):
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    def iterate_agency_date_day(self, form_data, meta_values, subevs, quotas, orders, checkins):
        yield [
                  self.meta_name,
                  "",
              ] + [
                  dt.strftime('%m/%d/%Y') for dt in self._date_iter()
              ]

        for mv in meta_values:
            events = sorted([e for e in self.cached_events if e.meta_data[self.meta_name] == mv], key=lambda e: str(e.name))

            yield [
                      mv, "Sum of Quotas",
                  ] + [
                      sum((quotas.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in self._date_iter()
                  ]
            yield [
                      "", "Sum of Orders",
                  ] + [
                      sum((orders.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in self._date_iter()
                  ]
            yield [
                      "", "Sum of Checked in",
                  ] + [
                      sum((checkins.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in self._date_iter()
                  ]
            yield []
        yield [
                  "Total", "Sum of Quotas",
              ] + [
                  sum((quotas.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in self._date_iter()
              ]
        yield [
                  "", "Sum of Orders",
              ] + [
                  sum((orders.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in self._date_iter()
              ]
        yield [
                  "", "Sum of Checked in",
              ] + [
                  sum((checkins.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in self._date_iter()
              ]

    def prepare_xlsx_sheet_agency_date_day(self, ws):
        ws.freeze_panes = 'C2'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        for i, dt in enumerate(self._date_iter()):
            ws.column_dimensions[get_column_letter(3 + i)].width = 15

    def iterate_agency_date_week(self, form_data, meta_values, subevs, quotas, orders, checkins):
        weeks = list(self._week_iter())
        yield [
                  "",
                  "",
              ] + [
                  "Week" for w in weeks
              ]
        yield [
                  "",
                  "First day",
              ] + [
                  w[0].strftime('%m/%d/%Y') for w in weeks
              ]
        yield [
                  self.meta_name,
                  "Last day",
              ] + [
                  w[-1].strftime('%m/%d/%Y') for w in weeks
              ]

        for mv in meta_values:
            events = sorted([e for e in self.cached_events if e.meta_data[self.meta_name] == mv], key=lambda e: str(e.name))

            yield [
                      mv, "Sum of Quotas",
                  ] + [
                      sum(sum((quotas.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in w) for w in weeks
                  ]
            yield [
                      "", "Sum of Orders",
                  ] + [
                      sum(sum((orders.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in w) for w in weeks
                  ]
            yield [
                      "", "Sum of Checked in",
                  ] + [
                      sum(sum((checkins.get((dt, e.pk), 0) or 0 for e in events), start=0) for dt in w) for w in weeks
                  ]
            yield []
        yield [
                  "Total", "Sum of Quotas",
              ] + [
                  sum(sum((quotas.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in w) for w in weeks
              ]
        yield [
                  "", "Sum of Orders",
              ] + [
                  sum(sum((orders.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in w) for w in weeks
              ]
        yield [
                  "", "Sum of Checked in",
              ] + [
                  sum(sum((checkins.get((dt, e.pk), 0) or 0 for e in self.cached_events), start=0) for dt in w) for w in weeks
              ]

    def prepare_xlsx_sheet_agency_date_week(self, ws):
        ws.freeze_panes = 'C4'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        for i, dt in enumerate(self._week_iter()):
            ws.column_dimensions[get_column_letter(3 + i)].width = 15


class CapacityCreationReport(BaseMSLE):
    identifier = 'capacity_creation'
    verbose_name = 'Capacity creation'
    meta_name = 'AgencyNumber'

    sheets = [
        ('date_agency', _('By date and agency')),
    ]

    @property
    def export_form_fields(self):
        defdate_start = now().astimezone(get_current_timezone()).date()
        defdate_end = now().astimezone(get_current_timezone()).date() + timedelta(days=6)

        f = OrderedDict(
            list(super().export_form_fields.items()) + [
                ('date_from',
                 forms.DateField(
                     label=_('Start date'),
                     widget=forms.DateInput(attrs={'class': 'datepickerfield'}),
                     initial=defdate_start,
                 )),
                ('date_to',
                 forms.DateField(
                     label=_('End date'),
                     widget=forms.DateInput(attrs={'class': 'datepickerfield'}),
                     initial=defdate_end,
                 )),
            ]
        )
        if self.is_multievent and self.events.first():
            organizer = self.events.first().organizer
            for mp in organizer.meta_properties.prefetch_related('event_values'):
                if mp.name != self.meta_name:
                    continue
                values = sorted(list({v.value for v in mp.event_values.all()}))
                f['meta:{}'.format(mp.name)] = forms.MultipleChoiceField(
                    label=mp.name,
                    choices=[(v, v) for v in values],
                    widget=forms.CheckboxSelectMultiple(
                        attrs={'class': 'scrolling-multiple-choice'}
                    ),
                    initial=values,
                )
        return f

    def iterate_sheet(self, form_data, sheet):
        if self.events.first():
            self.tz = self.events.first().timezone
        else:
            self.tz = pytz.UTC
        self.date_from = form_data['date_from']
        self.date_until = form_data['date_to']
        if isinstance(self.date_from, str):
            self.date_from = parse(self.date_from).date()
        self.datetime_from = make_aware(datetime.combine(
            self.date_from,
            time(hour=0, minute=0, second=0, microsecond=0)
        ), self.tz)

        if isinstance(self.date_until, str):
            self.date_until = parse(self.date_until).date()
        self.datetime_until = make_aware(datetime.combine(
            self.date_until + timedelta(days=1),
            time(hour=0, minute=0, second=0, microsecond=0)
        ), self.tz)

        if hasattr(self, 'iterate_' + sheet):
            yield from getattr(self, 'iterate_' + sheet)(form_data)

    def _date_iter(self):
        dt = self.date_from
        while dt <= self.date_until:
            yield dt
            dt += timedelta(days=1)

    def iterate_date_agency(self, form_data):
        yield [
            "Event created", self.meta_name, "Number of Events", "Sum of Quotas", "Sum of Orders", "Sum of Checked in"
        ]

        qs = self.events.annotate(
            creation_datetime=Subquery(
                LogEntry.objects.filter(
                    event=OuterRef('pk')
                ).order_by().values('event').annotate(c=Min('datetime')).values('c')
            ),
            meta_value=Subquery(
                EventMetaValue.objects.filter(
                    property__name=self.meta_name,
                    event=OuterRef('pk')
                ).order_by().values('event').annotate(c=Max('value')).values('c')
            ),
            n_quotas=Subquery(
                Quota.objects.filter(
                    size__isnull=False,
                    event=OuterRef('pk')
                ).order_by().values('event').annotate(s=Sum('size')).values('s')
            ),
            n_orders=Subquery(
                OrderPosition.objects.filter(
                    order__event=OuterRef('pk'),
                    order__status__in=(Order.STATUS_PAID, Order.STATUS_PENDING),
                ).order_by().values('order__event').annotate(s=Count('*')).values('s')
            ),
            n_checkins=Subquery(
                OrderPosition.objects.annotate(
                    has_checkin=Exists(Checkin.objects.filter(position=OuterRef('pk')))
                ).filter(
                    has_checkin=True,
                    order__event=OuterRef('pk'),
                    order__status__in=(Order.STATUS_PAID, Order.STATUS_PENDING),
                ).order_by().values('order__event').annotate(s=Count('*')).values('s')
            ),
        ).annotate(
            creation_date=TruncDay(F('creation_datetime'))
        ).filter(
            creation_datetime__gte=self.datetime_from,
            creation_datetime__lt=self.datetime_until,
            meta_value__in=form_data['meta:' + self.meta_name]
        ).order_by().values(
            'creation_date', 'meta_value'
        ).annotate(
            sum_events=Count('*'),
            sum_quotas=Sum('n_quotas'),
            sum_orders=Sum('n_orders'),
            sum_checkins=Sum('n_checkins'),
        ).order_by(
            'creation_date', 'meta_value'
        )

        for r in qs:
            yield [
                r['creation_date'].strftime('%m/%d/%Y'),
                r['meta_value'],
                r['sum_events'] or 0,
                r['sum_quotas'] or 0,
                r['sum_orders'] or 0,
                r['sum_checkins'] or 0,
            ]

    def prepare_xlsx_sheet_date_agency(self, ws):
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
